from openpyxl import load_workbook

def get():
    if text == '1':
        bot.register_next_step_handler(msg, answer3)
        
        
        
        filename = 'my_project.xlsx'
        wb = load_workbook(filename)
        sheet = wb.active
        id_ = 1
        bot.send_message(chat_id=msg.from_user.id, surname = 'Введите фамилию')
        bot.send_message(chat_id=msg.from_user.id, first_name ='Введите имя')
        bot.send_message(chat_id=msg.from_user.id, patronym = 'Введите отчество')
        bot.send_message(chat_id=msg.from_user.id, phone = 'Введите телефон')
        bot.send_message(chat_id=msg.from_user.id, comment = 'Введите комментарий')
        surname, first_name, patronym,phone,comment = map(int, msg.text.split()) 
        for row in range(sheet.max_row + 1, sheet.max_row + 2):        
            sheet[row][0].value = row - 1
            sheet[row][1].value = surname 
            sheet[row][2].value = first_name 
            sheet[row][3].value = patronym 
            sheet[row][4].value = phone
            sheet[row][5].value = comment
        wb.save(filename) 
        wb.close() 