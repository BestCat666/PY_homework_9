from telebot import TeleBot, types
import os
import pandas
import openpyxl
from openpyxl import load_workbook

os.chdir(os.path.dirname(__file__))
 
TOKEN = ''
bot = TeleBot(TOKEN)

@bot.message_handler(commands=['start', 'help'])
def answer(msg: types.Message):
    bot.send_message(chat_id=msg.from_user.id, text='Введите номер операции: 1 - добавление записи, 2 - вывод записи на экран, 3 - импорт, 4 - экспорт')


@bot.message_handler()
def answer(msg: types.Message):
    text = msg.text
    if text == '1':
        bot.register_next_step_handler(msg, answer1)
        bot.send_message(chat_id=msg.from_user.id, text='Введите фамилию, имя, отчество, номер телефона и комментраий через пробел')
    

    if text == '2':                     
        book = openpyxl.open('my_project.xlsx', 'r')
        sheet = book.active
        for row in range(1, sheet.max_row + 1):
            surname = sheet[row][1].value 
            first_name  = sheet[row][2].value
            patronym = sheet[row][3].value
            phone = sheet[row][4].value
            comment = sheet[row][5].value
            txt = f'{surname}, {first_name}, {patronym}, {phone}, {comment}'      
            bot.send_message(chat_id=msg.from_user.id, text = txt)
    
    # if text == '2':   #  вывод содержимого каждой ячейки отдельным сообщением                 
    #     book = openpyxl.open('my_project.xlsx', 'r')
    #     sheet = book.active
    #     for row in range(1, sheet.max_row + 1):
    #         for cell in row:
    #             bot.send_message(chat_id=msg.from_user.id, text = cell.value)
            

    if text == '3':
        bot.send_message(chat_id=msg.from_user.id, text='Добавьте документ для импортирования')
        bot.register_next_step_handler(msg, answer3)
    if text == '4':
        bot.send_document(chat_id=msg.from_user.id, document = open('my_project.xlsx', 'rb'))  


def answer1(msg):
    surname, first_name, patronym, phone, comment = map(str, msg.text.split())
    filename = 'my_project.xlsx'
    wb = load_workbook(filename)
    sheet = wb.active
    id_ = 0
    for row in range(sheet.max_row + 1, sheet.max_row + 2):        
        sheet[row][0].value = row - 2
        sheet[row][1].value = surname 
        sheet[row][2].value = first_name 
        sheet[row][3].value = patronym 
        sheet[row][4].value = phone
        sheet[row][5].value = comment
    wb.save(filename) 
    wb.close() 
    bot.send_message(chat_id=msg.from_user.id, text='Запись добавлена')


@bot.message_handler(content_types=['document'])
def answer3(msg: types.Message):
    filename = msg.document.file_name
    with open(filename, 'wb') as file:
        file.write(bot.download_file(bot.get_file(msg.document.file_id).file_path))
    bot.send_message(chat_id=msg.from_user.id, text='Документ импортирован')
    d1 = pandas.read_excel(filename)
    d2 = pandas.read_excel('my_project.xlsx')
    d3 = pandas.concat([d2, d1], axis = 0, ignore_index = True)
    d3 = d3.reset_index(inplace = False, drop = True)
    d3['ID'] = d3.index
    d3 = d3.set_index('ID')
    d3.to_excel('my_project.xlsx')
    
bot.polling()

