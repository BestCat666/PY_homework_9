import openpyxl
import pandas
# Вариант №1

def read():
    book = openpyxl.open('my_project.xlsx', 'r')
    sheet = book.active
    for row in range(sheet.max_row + 1, sheet.max_row + 2):
    # for row in sheet.iter_rows():
    #     for cell in row:
            print(cell.value, end=" ")
        print()
read()
# Вариант №2
# def read():
#     data = pandas.read_excel("my_project.xlsx")
#     print(data)